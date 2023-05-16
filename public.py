from colorama import Fore
import pandas as pd
import openpyxl
from openpyxl.workbook import Workbook
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Side, Border
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

def log_output(log_level, message):
    if log_level == 'ERROR':
        # 将错误日志以红色打印
        print(f"{Fore.RED}log_level=[ERROR] {message}")
    elif log_level == 'INFO':
        # 将信息日志以绿色打印
        print(f"{Fore.GREEN}log_level=[INFO] {message}")
    else:
        # 默认情况下，将日志以白色打印
        print(f"{Fore.WHITE}log_level={log_level} {message}")


def sort_key(element):
    # 指定用于排序的字典键
    key_to_sort = 'your_key'
    # 获取指定键的值
    value = element.get(key_to_sort)
    # 在这里可以进行一些自定义的处理，如将值转换为整数等
    return value


import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# 你的字典列表



def write_to_excel(list_of_dicts, excel_path):
    # 将列表转换为pandas DataFrame
    df = pd.DataFrame(list_of_dicts)

    # 对DataFrame进行排序，确保"布线类型"在第一列
    df = df.set_index('布线类型').reset_index()

    # 将DataFrame保存为excel文件
    df.to_excel(excel_path, index=False)

    # 使用openpyxl加载excel文件，对第一列进行合并
    wb = load_workbook(excel_path)
    ws = wb.active

    last_cell = None
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        cell = row[0]
        if last_cell is None:
            last_cell = cell
        elif last_cell.value == cell.value:
            ws.merge_cells(start_row=last_cell.row, start_column=last_cell.column, end_row=cell.row, end_column=cell.column)
            ws.cell(row=last_cell.row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        else:
            last_cell = cell

    # 保存更改
    wb.save(excel_path)



import openpyxl
from openpyxl.styles import Font, Alignment
def write_list_to_excel_new(data_list, output_file, sheet_name):
    df = pd.DataFrame(data_list)

    # 创建一个新的Workbook
    wb = openpyxl.load_workbook(output_file)
    ws = wb.create_sheet(title=sheet_name)
    # 设置列宽
    column_widths = [20, 10, 10, 10, 10, 10, 15, 10, 30, 10, 20, 10, 10, 10, 10, 10, 15, 10, 30, 10, 10, 10, 30, 30, 30, 30, 30, 30]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = width / 1  # 将像素宽度转换为字符宽度

    # 将数据框转换为行并将其添加到工作表中
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # 设置所有单元格的对齐方式
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # 设置第一行的背景色和字体颜色
            if r_idx == 1:
                cell.fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
                cell.font = Font(color='FFFFFF')

            # 设置C列从第二行开始左对齐
            if c_idx == 3 and r_idx > 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')



    font = Font(name='微软雅黑', size=10)

    # 创建边框对象
    side = Side(style='thin', color='000000')
    border = Border(left=side, right=side, top=side, bottom=side)

    # 遍历工作表中的所有单元格
    for row in ws.iter_rows():
        for cell in row:
            # 应用字体
            cell.font = font

            # 应用边框
            cell.border = border

    # 合并A列单元格相同的值
    # previous_value = None
    # start_row = 1
    # for row in range(1, ws.max_row + 1):
    #     current_value = ws.cell(row=row, column=1).value
    #     if previous_value is None:
    #         previous_value = current_value
    #     elif current_value != previous_value:
    #         if row - 1 != start_row:
    #             ws.merge_cells(start_row=start_row, end_row=row - 1, start_column=1, end_column=1)
    #         previous_value = current_value
    #         start_row = row
    #
    # if ws.max_row != start_row:
    #     ws.merge_cells(start_row=start_row, end_row=ws.max_row, start_column=1, end_column=1)


    # 创建字体对象，设置字体颜色为白色
    header_font = Font(name='微软雅黑', size=10, color='FFFFFF')

    # 创建填充对象，设置填充颜色为红色
    header_fill = PatternFill(fill_type='solid', fgColor='FF0000')

    # 遍历第一行的单元格
    for cell in ws[1]:
        # 应用字体
        cell.font = header_font

        # 应用填充
        cell.fill = header_fill

    # 从第二行开始，设置E列和G列的单元格为右对齐
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')  # E列
        ws.cell(row=row, column=7).alignment = Alignment(horizontal='right')  # G列
    ws.sheet_view.zoomScale = 175
    # 保存工作簿
    wb.save(output_file)